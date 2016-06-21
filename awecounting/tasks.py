from __future__ import absolute_import

import datetime
from django.conf import settings
import xlrd
from openpyxl.workbook import Workbook as openpyxlWorkbook
from openpyxl import load_workbook
from django.core.mail import send_mail
from django.db import transaction
from awecounting.utils.helpers import empty_to_zero, zero_for_none
from .celery import app
from apps.inventory.models import ItemCategory, Unit, InventoryAccount, Item, set_transactions
from apps.ledger.models import Party


def xls_to_xlsx(content):
    xlsBook = xlrd.open_workbook(file_contents=content.read())
    workbook = openpyxlWorkbook()

    for i in xrange(0, xlsBook.nsheets):
        xlsSheet = xlsBook.sheet_by_index(i)
        sheet = workbook.active if i == 0 else workbook.create_sheet()
        sheet.title = xlsSheet.name

        for row in xrange(0, xlsSheet.nrows):
            for col in xrange(0, xlsSheet.ncols):
                sheet.cell(row=row + 1, column=col + 1).value = xlsSheet.cell_value(row, col)

    return workbook


@app.task
def stock_tally(fl, company, user):
    from apps.haul.views import xls_stock_tally

    if fl.name.endswith('.xls'):
        wb = xls_to_xlsx(fl)
    else:
        wb = load_workbook(fl)
    sheets = wb.worksheets
    inventory_account_no = InventoryAccount.get_next_account_no(company=company)
    cnt = 0
    for sheet in sheets:
        category, category_created = ItemCategory.objects.get_or_create(name=sheet.title,
                                                                        company=company)
        rows = tuple(sheet.iter_rows())
        with transaction.atomic():
            unit, created = Unit.objects.get_or_create(name="Pieces", company=company)
            for row in rows[5:]:
                params = xls_stock_tally(row)
                if params.get('particulars') not in ['Grand Total', 'Total', 'total']:
                    rate = empty_to_zero(params.get('rate'))
                    quantity = empty_to_zero(params.get('quantity'))
                    item = Item(name=params.get('particulars'), cost_price=rate, category=category,
                                unit=unit, company=company, oem_no=empty_to_zero(params.get('oem_number')))
                    cnt += 1
                    item.save(account_no=inventory_account_no)
                    inventory_account_no += 1
                    if quantity != 0:
                        set_transactions(item.account, datetime.date.today(),
                                         ['dr', item.account, quantity])
    send_mail('Inventory Stock Import complete', str(cnt) + ' inventory records imported.', settings.DEFAULT_FROM_EMAIL,
              [user.email], fail_silently=False)


@app.task
def debtor_tally(fl, company, post, user):
    from apps.haul.views import xls_debtor_tally

    if fl.name.endswith('.xls'):
        wb = xls_to_xlsx(fl)
    else:
        wb = load_workbook(fl)
    sheets = wb.worksheets
    cnt = 0
    for sheet in sheets:
        rows = tuple(sheet.iter_rows())
        with transaction.atomic():
            for row in rows[5:]:
                params = xls_debtor_tally(row)
                if type(params.get('debit')) != str and type(params.get('debit')) != str:
                    if 'new_party' in post:
                        party = Party.objects.create(name=params.get('particulars'), company=company)
                    else:
                        party, party_created = Party.objects.get_or_create(name=params.get('particulars'),
                                                                           company=company)
                    cnt += 1
                    party.customer_account.opening_cr = zero_for_none(params.get('credit'))
                    party.customer_account.opening_dr = zero_for_none(params.get('debit'))
                    party.customer_account.save()
                    party.save()
    send_mail('Import complete', str(cnt) + ' debtors imported.', settings.DEFAULT_FROM_EMAIL, [user.email], fail_silently=False)
