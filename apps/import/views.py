import re
import datetime

from django.core.urlresolvers import reverse

from django.http import HttpResponseRedirect

from django.shortcuts import render

from awecounting.utils.helpers import zero_for_none, empty_to_zero
from ..ledger.models import Party
from ..inventory.models import ItemCategory, Unit, InventoryAccount, Item, set_transactions
from forms import ImportFile
from openpyxl import load_workbook
import xlrd
from openpyxl.workbook import Workbook as openpyxlWorkbook


def xls_to_xlsx(content):
    xlsBook = xlrd.open_workbook(file_contents=content.read())
    workbook = openpyxlWorkbook()

    for i in xrange(0, xlsBook.nsheets):
        xlsSheet = xlsBook.sheet_by_index(i)
        sheet = workbook.active if i == 0 else workbook.create_sheet()
        sheet.title = xlsSheet.name

        for row in xrange(0, xlsSheet.nrows):
            for col in xrange(0, xlsSheet.ncols):
                sheet.cell(row=row + 1, column=col + 1).value = xlsSheet.cell_value(row, col)

    return workbook


def xls_debtor_tally(row):
    dct = {}
    header = {'A': 'particulars', 'B': 'debit', 'C': 'credit'}
    for cell in row:
        dct[header.get(cell.column)] = cell.value
    return dct


def xls_stock_tally(row):
    dct = {}
    header = {'A': 'particulars', 'B': 'quantity', 'C': 'rate', 'D': 'value'}
    for cell in row:
        dct[header.get(cell.column)] = cell.value
        if cell.column == "A":
            data = re.search("\(([^\)]+)\)", cell.value)
            if data:
                dct['oem_number'] = data.group()[1:-1]
    return dct


def import_debtor_tally(request):
    if request.POST:
        form = ImportFile(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            if file.name.endswith('.xls'):
                wb = xls_to_xlsx(file)
            else:
                wb = load_workbook(file)
            sheet = wb.worksheets[0]
            rows = tuple(sheet.iter_rows())
            for row in rows[5:]:
                params = xls_debtor_tally(row)
                if type(params.get('debit')) != str and type(params.get('debit')) != str:
                    if 'new_party' in request.POST:
                        party = Party.objects.create(name=params.get('particulars'), company=request.company)
                    else:
                        party, party_created = Party.objects.get_or_create(name=params.get('particulars'),
                                                                           company=request.company)
                    party.customer_account.opening_cr = zero_for_none(params.get('credit'))
                    party.customer_account.opening_dr = zero_for_none(params.get('debit'))
                    party.customer_account.save()
                    party.save()
            return HttpResponseRedirect(reverse('party_list'))
    form = ImportFile()
    return render(request, 'import/import_debtor_tally.html', {'form': form})


def import_stock_tally(request):
    if request.POST:
        form = ImportFile(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            if file.name.endswith('.xls'):
                wb = xls_to_xlsx(file)
            else:
                wb = load_workbook(file)
            sheets = wb.worksheets
            for sheet in sheets:
                category, category_created = ItemCategory.objects.get_or_create(name=sheet.title,
                                                                                company=request.company)
                rows = tuple(sheet.iter_rows())
                unit, created = Unit.objects.get_or_create(name="Pieces", company=request.company)
                account_no = InventoryAccount.get_next_account_no(company=request.company)
                for row in rows[5:]:
                    params = xls_stock_tally(row)
                    if params.get('particulars') not in ['Grand Total', 'Total', 'total']:
                        rate = params.get('rate')
                        quantity = params.get('quantity')
                        if rate == '':
                            rate = empty_to_zero(rate)
                        if quantity == '':
                            quantity = empty_to_zero(quantity)

                        item = Item(name=params.get('particulars'), cost_price=zero_for_none(rate), category=category,
                                    unit=unit, company=request.company)
                        if params.get('oem_number'):
                            item.oem_no = params.get('oem_number')
                        item.save(account_no=account_no)
                        account_no = account_no + 1
                        item.account.current_balance = zero_for_none(quantity)
                        item.account.save()
                        if quantity > 0:
                            set_transactions(item.account, datetime.date.today(),
                                             ['dr', item.account, quantity])
            return HttpResponseRedirect(reverse('item_list'))
    form = ImportFile()
    return render(request, 'import/import_stock_tally.html', {'form': form})
